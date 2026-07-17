#!/usr/bin/env python3
"""
Phase 1 · Build source pool for the July 16 IP webinar campaign.

Python version — bypasses the node_modules / symlink mess on the Mac side.
Uses only stdlib + openpyxl (already in anaconda base). Talks to Neon via
its HTTP endpoint so no postgres driver needed.

Unions three sources into one deduped XLSX:
  - investors table (Neon)
  - crm_entries table (Neon)
  - anker-network-*.xlsx (LinkedIn export)

Dedupe order: linkedin_url → normalized(name+firm) → email.

Usage:
  NEON_DATABASE_URL='postgresql://...' \\
    python3 scripts/oneshot/02-build-source-pool.py \\
      ~/Downloads/anker-network-2026-07-11.xlsx
"""
import json, os, re, sys
from urllib import request as urlrequest
from urllib.parse import urlparse
from openpyxl import Workbook, load_workbook

# ---------- Neon HTTP ----------
NEON_URL = os.environ.get("NEON_DATABASE_URL")
if not NEON_URL:
    sys.exit("NEON_DATABASE_URL missing")

parsed = urlparse(NEON_URL)
NEON_HOST = parsed.hostname
NEON_ENDPOINT = f"https://{NEON_HOST}/sql"

def sql(query, params=None):
    """POST a SQL query to Neon's HTTP endpoint. Returns list of dict rows."""
    body = json.dumps({"query": query, "params": params or []}).encode("utf-8")
    req = urlrequest.Request(
        NEON_ENDPOINT,
        data=body,
        headers={
            "Neon-Connection-String": NEON_URL,
            "Neon-Raw-Text-Output": "true",
            "Neon-Array-Mode": "true",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    with urlrequest.urlopen(req, timeout=60) as r:
        data = json.loads(r.read())
    fields = [f["name"] for f in data.get("fields", [])]
    rows = data.get("rows", [])
    return [dict(zip(fields, row)) for row in rows]

# ---------- Input ----------
if len(sys.argv) < 2:
    sys.exit("Usage: python 02-build-source-pool.py <path/to/anker-network-*.xlsx>")
network_xlsx = os.path.expanduser(sys.argv[1])

def norm(s):
    return (s or "").strip().lower()

def norm_url(s):
    return re.sub(r"^https?://(www\.)?", "", norm(s)).rstrip("/")

# ---------- Pull investors ----------
print("→ Pulling investors from Neon…", flush=True)
investors = sql("""
    SELECT id::text, name, email, linkedin_url, title, firm, tags,
           COALESCE(metadata->>'lp_type', 'investor')  AS lp_type,
           COALESCE(metadata->>'sectors', '')          AS sectors,
           COALESCE(metadata->>'location', '')         AS location
    FROM investors
    WHERE name IS NOT NULL
""")
print(f"  {len(investors)} rows")

# ---------- Pull crm_entries ----------
print("→ Pulling crm_entries from Neon…", flush=True)
crm = sql("""
    SELECT id::text,
           COALESCE(display_name, linkedin_data->>'name') AS name,
           display_email                                  AS email,
           display_linkedin                               AS linkedin_url,
           linkedin_data->>'title'                        AS title,
           linkedin_data->>'company'                      AS firm,
           linkedin_data->>'headline'                     AS headline,
           stage
    FROM crm_entries
    WHERE display_name IS NOT NULL OR linkedin_data IS NOT NULL
""")
print(f"  {len(crm)} rows")

# ---------- Read network XLSX ----------
print(f"→ Reading LinkedIn network export: {network_xlsx}", flush=True)
wb = load_workbook(network_xlsx, read_only=True, data_only=True)
ws = wb.worksheets[0]
headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
idx = {h: i for i, h in enumerate(headers)}
def h(row, key):
    i = idx.get(key)
    return row[i] if i is not None else None

network = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if not any(row):
        continue
    network.append({
        "name": h(row, "name"),
        "linkedin_url": h(row, "linkedin_url"),
        "headline": h(row, "headline"),
        "firm": h(row, "company"),
        "title": h(row, "title"),
        "location": h(row, "location"),
        "degree": h(row, "degree"),
        "previous_firm": h(row, "previous_company"),
        "previous_title": h(row, "previous_title"),
    })
wb.close()
print(f"  {len(network)} rows")

# ---------- Dedupe & merge ----------
print("→ Deduping (linkedin_url → name+firm → email)…", flush=True)
pool = {}
def upsert(row, source):
    key = None
    if row.get("linkedin_url"):
        key = f"li:{norm_url(row['linkedin_url'])}"
    elif row.get("name"):
        key = f"n:{norm(row['name'])}|f:{norm(row.get('firm'))}"
    elif row.get("email"):
        key = f"e:{norm(row['email'])}"
    else:
        return
    existing = pool.get(key, {})
    merged = {**existing}
    for k, v in row.items():
        if v:
            merged[k] = v
    srcs = set(existing.get("sources", []))
    srcs.add(source)
    merged["sources"] = sorted(srcs)
    pool[key] = merged

for r in investors: upsert(r, "neon-investors")
for r in crm:       upsert(r, "neon-crm")
for r in network:   upsert(r, "linkedin-export")

rows = list(pool.values())
print(f"  {len(rows)} unique contacts after dedupe", flush=True)

# ---------- Write source-pool.xlsx ----------
print("→ Writing source-pool.xlsx…", flush=True)
out = Workbook()
out_ws = out.active
out_ws.title = "Pool"
cols = ["name", "email", "linkedin_url", "firm", "title", "headline",
        "lp_type", "sectors", "location", "tags", "degree",
        "previous_firm", "previous_title", "sources", "stage", "id"]
out_ws.append(cols)
for r in rows:
    vals = []
    for c in cols:
        v = r.get(c)
        if isinstance(v, (list, tuple)):
            v = ", ".join(str(x) for x in v)
        vals.append(v if v is not None else "")
    out_ws.append(vals)
# Bold header
for cell in out_ws[1]:
    cell.font = cell.font.copy(bold=True)
out.save("source-pool.xlsx")

# ---------- Coverage report ----------
c_both = c_email = c_li = c_none = 0
for r in rows:
    e, l = bool(r.get("email")), bool(r.get("linkedin_url"))
    if e and l:   c_both += 1
    elif e:       c_email += 1
    elif l:       c_li += 1
    else:         c_none += 1
print(f"\n✓ source-pool.xlsx written ({len(rows)} rows)")
print(f"Coverage:")
print(f"  email + linkedin: {c_both}")
print(f"  email only:       {c_email}")
print(f"  linkedin only:    {c_li}")
print(f"  neither:          {c_none}")
