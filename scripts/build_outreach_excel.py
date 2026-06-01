#!/usr/bin/env python3
"""
Summit Venture Studio Fund II — Outreach Excel Builder

Reads pipeline JSON (output of outreachPipeline.ts) and produces a
formatted .xlsx with 6 sheets:

  1. Enriched Profiles   — full per-LP enrichment data
  2. Email Drafts        — subject + body per LP
  3. LinkedIn DMs        — condensed ≤300-char DMs
  4. Campaign Summary    — stats + LP-type breakdown
  5. Multi-Touch Tracker — same-firm contacts
  6. Methodology         — pipeline explanation + voice guide

Formatting:
  - Navy headers          #1A3A5C (fills) + white bold text
  - Teal enrichment cols  #185FA5 (secondary header fill)
  - Alternating rows      white / #EBF3FB
  - Frozen panes (row 1)
  - Auto-filter on data sheets
  - Auto-width columns (capped at 65)
  - Colour-coded tab colours

Usage:
  python3 scripts/build_outreach_excel.py <pipeline_output.json> [output.xlsx]

  If no output path is given, writes to reports/svs-campaign-<timestamp>.xlsx
"""

import json
import sys
import os
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# ── Colour constants ──────────────────────────────────────────────────────────

NAVY          = "1A3A5C"   # primary header fill
TEAL          = "185FA5"   # enrichment-column header fill
WHITE         = "FFFFFF"
ALT_ROW       = "EBF3FB"   # alternating row fill
ACCENT_GREEN  = "217346"
ACCENT_ORANGE = "C55A11"
ACCENT_PURPLE = "7030A0"
LI_BLUE       = "0A66C2"
HEADER_FONT_SZ = 10
DATA_FONT_SZ   = 9

# ── Helpers ───────────────────────────────────────────────────────────────────

def hfill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def hfont(bold=False, color=WHITE, size=HEADER_FONT_SZ) -> Font:
    return Font(bold=bold, color=color, size=size, name="Calibri")

def dfont(bold=False, size=DATA_FONT_SZ) -> Font:
    return Font(bold=bold, size=size, name="Calibri")

def center_align(wrap=False) -> Alignment:
    return Alignment(horizontal="center", vertical="top", wrap_text=wrap)

def left_align(wrap=False) -> Alignment:
    return Alignment(horizontal="left", vertical="top", wrap_text=wrap)

def thin_border() -> Border:
    s = Side(style="thin", color="D0D7E0")
    return Border(bottom=s)

def apply_header_row(ws, headers: list[str], fills: list[str] | None = None):
    """Apply navy (or per-column) fill + white bold text to row 1."""
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        fill_color = fills[col_idx - 1] if fills else NAVY
        cell.fill = hfill(fill_color)
        cell.font = hfont(bold=True)
        cell.alignment = center_align(wrap=False)
        cell.border = thin_border()

def apply_data_rows(ws, data: list[list], wrap_cols: set[int] | None = None):
    """Write data rows with alternating fills."""
    wrap_cols = wrap_cols or set()
    for row_idx, row in enumerate(data, start=2):
        fill = hfill(ALT_ROW if row_idx % 2 == 0 else WHITE)
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value if value is not None else ""
            cell.fill = fill
            cell.font = dfont()
            cell.alignment = left_align(wrap=col_idx in wrap_cols)

def auto_width(ws, headers: list[str], max_width=65):
    """Set column widths based on content."""
    widths = [min(len(h) + 4, max_width) for h in headers]
    for row in ws.iter_rows(min_row=2, values_only=True):
        for i, val in enumerate(row):
            if val is None:
                continue
            first_line = str(val).split("\n")[0]
            widths[i] = min(max(widths[i], len(first_line) + 2), max_width)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def freeze_and_filter(ws, ref: str):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ref

def set_tab_color(ws, hex_color: str):
    ws.sheet_properties.tabColor = hex_color

# ── Sheet 1: Enriched Profiles ────────────────────────────────────────────────

ENRICHED_HEADERS = [
    "#", "Tier", "Score", "Name", "Title/Role", "LP Type", "Tags",
    "Location", "Email", "LinkedIn", "Sectors",
    "Why This Contact",          # original
    "Firm Intelligence",         # enriched ← teal
    "Investment Mandate",        # enriched ← teal
    "Personalisation Hook",      # enriched ← teal
    "Multi-Touch Note",          # enriched ← teal
    "Batch", "Primary Channel", "Outreach Status",
]

# Columns that are "enriched" (0-based indices → teal fill in header)
ENRICHED_TEAL_COLS = {12, 13, 14, 15}  # Firm Intel, Mandate, Hook, Multi-Touch

def build_enriched_sheet(ws, enriched: list[dict]):
    fills = [
        TEAL if i in ENRICHED_TEAL_COLS else NAVY
        for i in range(len(ENRICHED_HEADERS))
    ]
    apply_header_row(ws, ENRICHED_HEADERS, fills)

    rows = []
    for p in enriched:
        multi = ""
        if p.get("isMultiTouch"):
            multi = f"Follow-up: prior contact at firm is {p.get('multiTouchPriorContact', '')}"
        rows.append([
            p.get("id"), p.get("tier"), p.get("score"),
            p.get("name"), p.get("titleRole"), p.get("lpType"),
            p.get("tags"), p.get("location"), p.get("email"),
            p.get("linkedin"), p.get("sectors"), p.get("whyThisContact"),
            p.get("firmIntelligence"), p.get("investmentMandate"),
            p.get("personalisationHook"), multi,
            p.get("batch"), p.get("primaryChannel", ""),
            p.get("outreachStatus", "Draft"),
        ])

    apply_data_rows(ws, rows, wrap_cols={12, 13, 14, 15, 16})
    auto_width(ws, ENRICHED_HEADERS)
    freeze_and_filter(ws, f"A1:{get_column_letter(len(ENRICHED_HEADERS))}1")
    # Row heights
    for row in ws.iter_rows(min_row=2):
        ws.row_dimensions[row[0].row].height = 64
    set_tab_color(ws, NAVY)

# ── Sheet 2: Email Drafts ─────────────────────────────────────────────────────

EMAIL_HEADERS = [
    "#", "Name", "LP Type", "Email",
    "Subject", "Body",
    "Primary Channel", "Voice Notes", "Status",
]

def build_email_sheet(ws, drafts: list[dict]):
    apply_header_row(ws, EMAIL_HEADERS)
    rows = [
        [
            d.get("investorId"), d.get("name"), d.get("lpType"), d.get("email"),
            d.get("subject"), d.get("body"),
            d.get("primaryChannel"), d.get("voiceNotes"), d.get("outreachStatus", "Draft"),
        ]
        for d in drafts
    ]
    apply_data_rows(ws, rows, wrap_cols={5, 6})
    auto_width(ws, EMAIL_HEADERS, max_width=80)
    freeze_and_filter(ws, f"A1:{get_column_letter(len(EMAIL_HEADERS))}1")
    for row in ws.iter_rows(min_row=2):
        ws.row_dimensions[row[0].row].height = 120
    set_tab_color(ws, TEAL)

# ── Sheet 3: LinkedIn DMs ─────────────────────────────────────────────────────

LI_HEADERS = [
    "#", "Name", "LP Type", "LinkedIn URL", "DM (first touch)", "Chars", "Voice Notes",
]

def build_linkedin_sheet(ws, enriched: list[dict], drafts: list[dict]):
    draft_map = {d["investorId"]: d for d in drafts}
    apply_header_row(ws, LI_HEADERS)
    rows = []
    for p in enriched:
        d = draft_map.get(p["id"], {})
        dm = d.get("linkedInDM", "")
        rows.append([
            p.get("id"), p.get("name"), p.get("lpType"),
            p.get("linkedin"), dm, len(dm), d.get("voiceNotes", ""),
        ])
    apply_data_rows(ws, rows, wrap_cols={4})
    auto_width(ws, LI_HEADERS, max_width=70)
    freeze_and_filter(ws, f"A1:{get_column_letter(len(LI_HEADERS))}1")
    for row in ws.iter_rows(min_row=2):
        ws.row_dimensions[row[0].row].height = 72
    set_tab_color(ws, LI_BLUE)

# ── Sheet 4: Campaign Summary ─────────────────────────────────────────────────

def build_summary_sheet(ws, stats: dict, meta: dict):
    title_font = Font(bold=True, size=14, color=NAVY, name="Calibri")
    section_font = Font(bold=True, size=10, color=WHITE, name="Calibri")

    def write(row, col, val, font=None, fill=None, align=None):
        cell = ws.cell(row=row, column=col, value=val)
        if font: cell.font = font
        if fill: cell.fill = fill
        if align: cell.alignment = align
        return cell

    write(1, 1, "Summit Venture Studio Fund II — Campaign Summary", title_font)
    ws.merge_cells("A1:D1")

    write(3, 1, "Generated", dfont(bold=True))
    write(3, 2, meta.get("generatedAt", "")[:10])
    write(4, 1, "Total LPs", dfont(bold=True))
    write(4, 2, stats.get("total", 0))
    write(5, 1, "Enrichment Batches", dfont(bold=True))
    write(5, 2, stats.get("batches", 0))
    write(6, 1, "Multi-Touch Pairs", dfont(bold=True))
    write(6, 2, stats.get("multiTouchCount", 0))
    write(7, 1, "Avg Match Score", dfont(bold=True))
    write(7, 2, stats.get("avgScore", 0))

    # LP type breakdown table
    row = 9
    write(row, 1, "LP Type", section_font, hfill(NAVY), center_align())
    write(row, 2, "Count", section_font, hfill(NAVY), center_align())
    write(row, 3, "% of Total", section_font, hfill(NAVY), center_align())
    row += 1
    total = stats.get("total", 1)
    for lp_type, count in sorted(stats.get("byLPType", {}).items(), key=lambda x: -x[1]):
        fill = hfill(ALT_ROW if row % 2 == 0 else WHITE)
        write(row, 1, lp_type, dfont(), fill)
        write(row, 2, count, dfont(), fill)
        write(row, 3, f"{round(count/total*100)}%", dfont(), fill)
        row += 1

    # Channel breakdown
    row += 1
    write(row, 1, "Channel", section_font, hfill(TEAL), center_align())
    write(row, 2, "Count", section_font, hfill(TEAL), center_align())
    row += 1
    for channel, count in stats.get("byChannel", {}).items():
        fill = hfill(ALT_ROW if row % 2 == 0 else WHITE)
        write(row, 1, channel, dfont(), fill)
        write(row, 2, count, dfont(), fill)
        row += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14
    set_tab_color(ws, ACCENT_GREEN)

# ── Sheet 5: Multi-Touch Tracker ──────────────────────────────────────────────

MT_HEADERS = [
    "Firm / Website", "Contact 1 Name", "Contact 1 Email",
    "Contact 2 Name", "Contact 2 Email", "Additional", "Note",
]

def build_multitouch_sheet(ws, enriched: list[dict]):
    apply_header_row(ws, MT_HEADERS)
    # Group by domain
    by_domain: dict[str, list] = {}
    for p in enriched:
        site = p.get("inferredWebsite", "") or p.get("linkedin", "")
        try:
            from urllib.parse import urlparse
            domain = urlparse(site if site.startswith("http") else f"https://{site}").hostname or site
            domain = domain.replace("www.", "")
        except Exception:
            domain = site
        if not domain:
            domain = p.get("name", "")[:20]
        by_domain.setdefault(domain, []).append(p)

    rows = []
    for domain, group in by_domain.items():
        if len(group) < 2:
            continue
        c1, c2, *rest = group
        rows.append([
            domain,
            c1.get("name"), c1.get("email"),
            c2.get("name"), c2.get("email"),
            ", ".join(r.get("name", "") for r in rest) if rest else "",
            "Coordinated outreach — reference prior contact in follow-up",
        ])

    if not rows:
        rows = [["No multi-touch pairs detected", "", "", "", "", "", ""]]

    apply_data_rows(ws, rows, wrap_cols={6})
    auto_width(ws, MT_HEADERS)
    freeze_and_filter(ws, f"A1:{get_column_letter(len(MT_HEADERS))}1")
    set_tab_color(ws, ACCENT_ORANGE)

# ── Sheet 6: Methodology ──────────────────────────────────────────────────────

def build_methodology_sheet(ws, brief: dict):
    title_font = Font(bold=True, size=14, color=NAVY, name="Calibri")
    section_font = Font(bold=True, size=10, color=WHITE, name="Calibri")
    body_font = Font(size=9, name="Calibri")

    def write(row, col, val, font=None, fill=None, align=None):
        cell = ws.cell(row=row, column=col, value=val)
        if font: cell.font = font
        if fill: cell.fill = fill
        if align: cell.alignment = align
        return cell

    write(1, 1, "Summit Venture Studio Fund II — Campaign Methodology", title_font)
    ws.merge_cells("A1:C1")

    steps = [
        ("Step 1", "Profile Enrichment", "AI research on each LP's firm mandate, investment focus, and a personalisation hook. Batched in groups of ≤10 to respect API rate limits."),
        ("Step 2", "Email Drafting", "Tone-matched email body + LinkedIn DM per LP type using Claude Sonnet. Multi-touch detection references prior firm contact."),
        ("Step 3", "Excel Export", "This workbook: 6 sheets covering enrichment, email drafts, LinkedIn DMs, summary, multi-touch, and methodology."),
        ("Step 4", "HTML Review UI", "Self-contained HTML file: expandable LP cards, type/score badges, filter pills, copy-to-clipboard email preview."),
    ]

    row = 3
    write(row, 1, "Pipeline Steps", section_font, hfill(NAVY))
    ws.merge_cells(f"A{row}:C{row}")
    row += 1
    for step, label, desc in steps:
        write(row, 1, step, dfont(bold=True), hfill(ALT_ROW if row % 2 == 0 else WHITE))
        write(row, 2, label, dfont(bold=True), hfill(ALT_ROW if row % 2 == 0 else WHITE))
        cell = write(row, 3, desc, body_font, hfill(ALT_ROW if row % 2 == 0 else WHITE))
        cell.alignment = left_align(wrap=True)
        ws.row_dimensions[row].height = 40
        row += 1

    row += 1
    write(row, 1, "Voice Principles", section_font, hfill(TEAL))
    ws.merge_cells(f"A{row}:C{row}")
    row += 1
    for i, principle in enumerate(brief.get("voicePrinciples", []), 1):
        fill = hfill(ALT_ROW if row % 2 == 0 else WHITE)
        write(row, 1, f"V{i}", dfont(bold=True), fill)
        cell = write(row, 2, principle, body_font, fill)
        cell.alignment = left_align(wrap=True)
        ws.merge_cells(f"B{row}:C{row}")
        ws.row_dimensions[row].height = 32
        row += 1

    row += 1
    write(row, 1, "Sender Brief", section_font, hfill(NAVY))
    ws.merge_cells(f"A{row}:C{row}")
    row += 1
    brief_rows = [
        ("Sender", brief.get("senderName")),
        ("Role", brief.get("senderRole")),
        ("Fund", brief.get("fundName")),
        ("Target", brief.get("fundTarget")),
        ("Minimum", brief.get("fundMinimum")),
        ("HQ", brief.get("fundHQ")),
        ("Close Target", "Q3 2026"),
        ("Contact", "invest@svsfund.vc"),
    ]
    for label, val in brief_rows:
        fill = hfill(ALT_ROW if row % 2 == 0 else WHITE)
        write(row, 1, label, dfont(bold=True), fill)
        write(row, 2, val or "", body_font, fill)
        row += 1

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 60
    set_tab_color(ws, ACCENT_PURPLE)

# ── Main ──────────────────────────────────────────────────────────────────────

def build_excel(pipeline_json_path: str, output_path: str | None = None):
    with open(pipeline_json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    enriched = data.get("enriched", [])
    drafts = data.get("drafts", [])
    stats = data.get("stats", {})
    brief = data.get("senderBrief", {})
    meta = {"generatedAt": data.get("generatedAt", datetime.utcnow().isoformat())}

    if not output_path:
        ts = datetime.utcnow().strftime("%Y-%m-%d_%H-%M-%S")
        os.makedirs("reports", exist_ok=True)
        output_path = f"reports/svs-campaign-{ts}.xlsx"

    wb = openpyxl.Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    ws1 = wb.create_sheet("Enriched Profiles")
    ws2 = wb.create_sheet("Email Drafts")
    ws3 = wb.create_sheet("LinkedIn DMs")
    ws4 = wb.create_sheet("Campaign Summary")
    ws5 = wb.create_sheet("Multi-Touch Tracker")
    ws6 = wb.create_sheet("Methodology")

    print("  Building Enriched Profiles sheet…")
    build_enriched_sheet(ws1, enriched)

    print("  Building Email Drafts sheet…")
    build_email_sheet(ws2, drafts)

    print("  Building LinkedIn DMs sheet…")
    build_linkedin_sheet(ws3, enriched, drafts)

    print("  Building Campaign Summary sheet…")
    build_summary_sheet(ws4, stats, meta)

    print("  Building Multi-Touch Tracker sheet…")
    build_multitouch_sheet(ws5, enriched)

    print("  Building Methodology sheet…")
    build_methodology_sheet(ws6, brief)

    wb.save(output_path)
    print(f"\n✅  Excel saved → {output_path}")
    return output_path


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python3 scripts/build_outreach_excel.py <pipeline_output.json> [output.xlsx]")
        sys.exit(1)

    json_path = sys.argv[1]
    out_path = sys.argv[2] if len(sys.argv) > 2 else None
    build_excel(json_path, out_path)
